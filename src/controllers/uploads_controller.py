from flask import jsonify, request, current_app
import os
import csv
import openpyxl

def upload_excel_controller():
    ALLOWED_EXTENSIONS = {"xlsx"}

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    # Validar que sea una solicitud de tipo multipart/form-data
    if not request.content_type.startswith("multipart/form-data"):
        return jsonify({"msg": "El Content-Type debe ser multipart/form-data", "ok": False}), 415

    # Validar que se incluya el archivo
    if 'file' not in request.files:
        return jsonify({"msg": "No se encontró un archivo en la solicitud", "ok": False}), 400

    # Obtener el archivo y otros datos del formulario
    file = request.files['file']

    if file.filename == '':
        return jsonify({"msg": "El archivo no tiene un nombre válido", "ok": False}), 400

    if not allowed_file(file.filename):
        return jsonify({"msg": f"Extensión no permitida. Extensiones válidas: {ALLOWED_EXTENSIONS}", "ok": False}), 400

    # Guardar el archivo en la ruta definida
    save_path_tmp = os.path.join(
        f"{current_app.config['UPLOAD_FOLDER']}/tmp", file.filename
    )
    file.save(save_path_tmp)
    
    save_path_csv = f"{current_app.config['UPLOAD_FOLDER']}/{file.filename.split('.')[0]}"
    os.makedirs(save_path_csv, exist_ok=True)
    
    try:
        wb = openpyxl.load_workbook(save_path_tmp)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            csv_filename = f"{sheet_name}.csv"
            
            with open(save_path_csv + "/" + csv_filename, mode='w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)

                # Escribir las filas en el CSV
                for row in sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)
            
        os.remove(save_path_tmp)
    except Exception as e:
        return jsonify({"msg": f"No se pudo convertir el archivo: {str(e)}", "ok": False}), 500

    
    return jsonify({ "ok": True })

    